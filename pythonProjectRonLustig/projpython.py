import requests
import openpyxl
import random
from bs4 import BeautifulSoup

w = []
#This method reads the top 1000 words in English and places them in a list
def read(w):
  url = "https://www.ef.edu/english-resources/english-vocabulary/top-1000-words/"
  response = requests.get(url)

  soup = BeautifulSoup(response.content, 'html.parser')

  s = soup.find('div', class_='content')

  lines = s.find_all('p')
    
  for line in lines:
     if 'a' in line:
        w.extend(line.text.split())
  return w

word=read(w)

#This method saves the top 100 words in English in Excel file
def save(word):
  wb = openpyxl.Workbook()
  ws = wb.active

  j = 1
  for i, word in enumerate(word):
    ws.cell(row=j, column=1).value = word
    j += 1
  # Save the workbook to an excel file
  wb.save("words.xlsx")
  return wb

wb1=save(word)

#This method reads a random word from the excel file 
def get_word(wb1):
  wb1 = openpyxl.load_workbook("words.xlsx")
  ws = wb1.active

  w1=[]
  for row in ws.values:
    for value in row:
     w1.append(value)

  rand_word = random.choice(w1)
  return rand_word


rand_word1=get_word(wb1)

#This method is responsible to update the display and place a character in place of a dash if user gueses the character correctly 
def update(dashes, char, words):
    for i in range(len(words)):
        if words[i] == char:
            dashes[i] = char
    return dashes


dashes = []
for letter in rand_word1:
    dashes.append('-')
  
print(" ".join(dashes))

#This last method has a while loop that will go through as long as user didn't make 5 mistakes or guessed the whole word correctly. 
# It also responsible to update the 
def check_win_lose(dashes,rand_word1):
  win = 0
  attempt = 0
  guessed_words = [] 

  while attempt < 5 and win != 1:

     user = input("Enter your guess: ")

     dis = update(dashes, user, rand_word1)
     print(" ".join(dis))

     if user not in rand_word1:
         attempt += 1

     if dis == list(rand_word1):
         win = 1
         print("You Won!")
         print('word was:',rand_word1)
         break

  if win == 0:
     print("You Lost!")
     print('word was:',rand_word1)

check_win_lose(dashes,rand_word1)